# Attendace Checking


   #importing openpyxl(module used to work on excel sheets)
from openpyxl import load_workbook     
   #importing option to colour
from openpyxl.styles import PatternFill   
   #importing date and time
from datetime import date                    

# -------------------------------------------------------------

f = open("text/excel_paths.txt")  # To get the path of the file


f_path = f.read()

   # Now to load the workbook in the path
wb = load_workbook(f_path)        


Date = date.today()               # To fetch today's date

   
ws = wb.active           # To choose an active worksheet   
ws.title = str(Date)           # To rename the sheet to the date

col = 2  # 1st row is Heading

data_str = ""  # To read from @zeta-bot

l=[]
i = 1
j=1
for i in range(1,6):
    for j in range (1,8):
        try:
            globals() [f'text.{i}.{j}'] = open(f'text.{i}.{j}.txt', 'r')          #TEXT FILE HAVING NAMES OF ALL STUDENTS
            globals() [f'var{i}.{j}'] = (globals() [f'text.{i}.{j}']).read()
            l.append(globals() [f'var{i}.{j}'])
        except FileNotFoundError:
            pass

print()
print("no errors")
print()

for i in l:
    data_str += i

#----------------------------------------------------------------------------------------
max_rows = ws.max_row

   #to traverse the rows in the excel file
for i in range(max_rows):

    element = ws["A" + str(col)].value  # A1 A2 A3 .....(column names)

    if element != None:
        if element.lower() in data_str.lower():        # presert condition
            ws["B" + str(col)] = "P"
            #colouring the sheet with green beside respective name
            ws["B" + str(col)].fill = PatternFill("solid", fgColor="71FF33")    

        
      else:                          # Abscent condition
            ws["B" + str(col)] = "A"   
            # To mark absentees with the colour red
            ws["B" + str(col)].fill = PatternFill("solid", fgColor="F50707")      
        col += 1

    # Now to save all the changes
wb.save(f_path)   
